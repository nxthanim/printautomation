#!/usr/bin/env python3
"""Generate the receipt_template.xlsx file used by tasks.py for receipt generation."""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


def generate_template(output_path: str = "receipt_template.xlsx") -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Receipt"

    thin = Side(style="thin")
    border = Border(top=thin, right=thin, bottom=thin, left=thin)
    header_font = Font(bold=True, size=11, name="Calibri")
    title_font = Font(bold=True, size=16, name="Calibri", color="1a73e8")
    value_font = Font(size=11, name="Calibri")
    header_fill = PatternFill(start_color="F0F4FF", end_color="F0F4FF", fill_type="solid")

    ws.merge_cells("A1:C1")
    cell = ws["A1"]
    cell.value = "PRINT AUTOMATION - OFFICIAL RECEIPT"
    cell.font = title_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:C2")
    ws["A2"].border = Border()

    labels = [
        ("A3", "TIN:"),
        ("A4", "Registration:"),
        ("A5", "Phone:"),
        ("A6", "Order ID:"),
        ("A7", "Customer:"),
        ("A8", "Date:"),
    ]

    for cell_ref, value in labels:
        cell = ws[cell_ref]
        cell.value = value
        cell.font = header_font
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.fill = header_fill
        cell.border = border

    for row in range(3, 9):
        for col in ["B", "C"]:
            cell = ws[f"{col}{row}"]
            cell.border = border
            cell.font = value_font
            cell.alignment = Alignment(horizontal="left", vertical="center")

        ws[f"B{row}"].border = Border()

    for col_letter in ["A", "B", "C"]:
        ws.column_dimensions[col_letter].width = 22

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 4
    ws.column_dimensions["C"].width = 36

    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    wb.save(output_path)
    print(f"Receipt template generated: {output_path}")
    return os.path.abspath(output_path)


if __name__ == "__main__":
    generate_template()
