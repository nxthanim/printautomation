import os
import logging
import subprocess
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from config import settings

logger = logging.getLogger("print_automation.tasks")

celery_app = None
if settings.celery_enabled:
    try:
        from celery import Celery
        celery_app = Celery(
            "print_automation",
            broker=settings.celery_broker,
            backend=settings.celery_backend,
        )
        celery_app.conf.update(
            task_serializer="json",
            accept_content=["json"],
            result_serializer="json",
            timezone="UTC",
            enable_utc=True,
            task_track_started=True,
            task_acks_late=True,
            worker_prefetch_multiplier=1,
        )
    except Exception as e:
        logger.warning("Celery not available, running tasks synchronously: %s", e)
        celery_app = None


def generate_receipt_sync(
    order_id: str,
    customer_name: str,
    customer_tin: str,
    customer_registration: str,
    customer_phone: str,
) -> str:
    receipt_path = os.path.join(settings.receipt_storage, f"receipt_{order_id}.xlsx")
    os.makedirs(settings.receipt_storage, exist_ok=True)
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Receipt"
        ws["A1"] = "OFFICIAL RECEIPT"
        ws["A3"] = "TIN No"
        ws["B3"] = customer_tin or ""
        ws["A4"] = "Reg No"
        ws["B4"] = customer_registration or ""
        ws["A5"] = "Phone"
        ws["B5"] = customer_phone or ""
        ws["A6"] = "Order"
        ws["B6"] = f"#{order_id[:8].upper()}"
        ws["A7"] = "Customer"
        ws["B7"] = customer_name
        ws["A8"] = "Date"
        ws["B8"] = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
        wb.save(receipt_path)
        logger.info("Receipt generated at %s", receipt_path)
        return receipt_path
    except Exception as exc:
        logger.error("Failed to generate receipt for order %s: %s", order_id, exc)
        raise


def send_email_sync(to: str, subject: str, body: str) -> str:
    import smtplib
    from email.mime.text import MIMEText
    try:
        msg = MIMEText(body)
        msg["Subject"] = subject
        msg["From"] = settings.notification_email
        msg["To"] = to
        with smtplib.SMTP(settings.smtp_host, settings.smtp_port) as s:
            if settings.smtp_user:
                s.starttls()
                s.login(settings.smtp_user, settings.smtp_password)
            s.send_message(msg)
        logger.info("Email sent to %s: %s", to, subject)
        return f"Email sent to {to}"
    except Exception as exc:
        logger.error("Email failed to %s: %s", to, exc)
        raise


def generate_receipt(order_id, customer_name, customer_tin, customer_registration, customer_phone):
    if celery_app:
        celery_app.send_task("generate_receipt", args=[order_id, customer_name, customer_tin, customer_registration, customer_phone])
    else:
        generate_receipt_sync(order_id, customer_name, customer_tin, customer_registration, customer_phone)


def send_email_notification(to, subject, body):
    if celery_app:
        celery_app.send_task("send_email_notification", args=[to, subject, body])
    else:
        send_email_sync(to, subject, body)
